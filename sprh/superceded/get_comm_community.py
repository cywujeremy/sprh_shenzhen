import requests
from bs4 import BeautifulSoup
import openpyxl as op
import time
import random
from urllib import request
from urllib.parse import quote
import string
from fake_useragent import UserAgent


def htmlIntoSoup(url):
    url = quote(url, safe=string.printable)
    time.sleep(random.random() * 10)
    headers = {
        'User-Agent': UserAgent(verify_ssl=False).random}
    proxy_list = [
        {'http': '115.29.170.58:8118'},
        {'http': '117.88.5.211:3000'},
        {'http': '222.95.240.228:3000'},
        {'http': '121.237.149.54:3000'},
        {'http': '117.88.5.126:3000'},
        {'http': '121.237.149.141:3000'},
        {'http': '222.95.241.123:3000'},
        {'http': '117.62.173.31:8118'},
        {'http': '117.88.5.15:3000'},
        {'http': '117.88.176.188:3000'},
        {'http': '60.190.250.120:8080'},
        {'http': '222.95.241.236:3000'},
        {'http': '222.95.144.46:3000'},
        {'http': '117.88.5.105:3000'},
        {'http': '222.95.144.72:3000'},
        {'http': '121.237.148.216:3000'},
        {'http': '117.88.4.91:3000'},
        {'http': '222.95.241.243:3000'},
        {'http': '106.122.205.36:8118'},
        {'http': '180.118.252.180:8118'},
        {'http': '121.237.149.158:3000'},
        {'http': '114.233.201.7:8118'},
        {'http': '117.87.180.144:8118'},
        {'http': '222.95.144.246:3000'},
        {'http': '117.88.177.86:3000'},
        {'http': '117.88.177.139:3000'},
        {'http': '121.237.148.195:3000'},
        {'http': '117.88.4.36:3000'},
        {'http': '117.88.4.55:3000'},
        {'http': '121.237.149.189:3000'},
        {'http': '117.88.176.176:3000'},
        {'http': '121.237.149.86:3000'},
        {'http': '117.85.166.121:8118'},
        {'http': '117.88.4.215:3000'},
        {'http': '121.237.149.180:3000'},
        {'http': '117.88.4.20:3000'},
        {'http': '117.88.176.213:3000'},
        {'http': '117.88.176.99:3000'},
        {'http': '121.237.149.62:3000'},
        {'http': '183.167.217.152:63000'},
        {'http': '117.88.5.167:3000'},
        {'http': '119.254.94.71:42788'},
        {'http': '117.88.176.169:3000'},
        {'http': '121.237.148.130:3000'},
        {'http': '121.237.148.228:3000'},
        {'http': '117.94.213.117:8118'},
        {'http': '121.237.149.16:3000'},
        {'http': '121.237.149.75:3000'},
        {'http': '121.237.148.151:3000'},
        {'http': '117.88.5.147:3000'},
        {'http': '117.88.177.3:3000'},
        {'http': '222.95.144.71:3000'},
        {'http': '14.212.249.10:8118'},

    ]
    proxy = random.choice(proxy_list)
    httpproxy_handler = request.ProxyHandler(proxy)
    opener = request.build_opener(httpproxy_handler)
    req = request.Request(url=url, headers=headers)
    response = opener.open(req)
    soup = BeautifulSoup(response.read(), 'lxml')
    return soup


def getUrl(soup):
    urlsoup = soup.find(class_='li-info').h3.contents
    try:
        sprh = urlsoup[3]['href']
    except TypeError:
        sprh = 'error'
    return sprh


def getINFO(soup):
    clip1 = soup.find(class_='comm-title')
    name = clip1.find(class_='sub-hd').previous_sibling
    clip2 = soup.find(class_='basic-parms-mod')
    fee = clip2.find(text='物业费：').parent.next_sibling.next_sibling.text
    return [name, fee]


def writeDistUrl():
    wb = op.load_workbook('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳商品房信息 - 副本.xlsx')
    ws1 = wb['NAME']
    ws2 = wb['URLs']
    i = ws2.max_row
    if i != 1:
        i = i + 1
    print(i)
    run_count = 0
    err_count = 0
    while i <= ws1.max_row:
        #name = ws1['A' + str(i)].value
        try:
            commurl = getUrl(htmlIntoSoup('https://shenzhen.anjuke.com/community/?kw='+ws1['A' + str(i)].value))
            ws2.append([commurl])
            print(ws1['A' + str(i)].value + '  URL查询成功！')
        except Exception as e:
            # ws2.append(['error'])
            print(ws1['A' + str(i)].value + '  URL查询错误！')
            print(e)
            err_count += 1
            if err_count >= 3:
                wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳商品房信息 - 副本.xlsx')
                time.sleep(60)
                i = i - 4
                err_count = 0
        i += 1
        run_count += 1
        if run_count >= 80:
            wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳商品房信息 - 副本.xlsx')
            run_count = 0
            time.sleep(30)
    wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳商品房信息 - 副本.xlsx')


def writeINFO():
    wb = op.load_workbook('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳商品房信息 - 副本.xlsx')
    ws1 = wb['URLs']
    ws2 = wb['INFO']
    i = ws2.max_row
    if i != 1:
        i = i + 1
    print(i)
    run_count = 0
    err_count = 0
    while i <= ws1.max_row:
        #name = ws1['A' + str(i)].value
        try:
            commurl = getINFO(htmlIntoSoup(ws1['A' + str(i)].value))
            ws2.append(commurl)
            print(commurl[0] + '  信息查询成功！')
        except Exception as e:
            # ws2.append(['error'])
            print(commurl[0] + '  信息查询错误！')
            print(e)
            err_count += 1
            if err_count >= 3:
                wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳商品房信息 - 副本.xlsx')
                time.sleep(60)
                i = i - 4
                err_count = 0
        i += 1
        run_count += 1
        if run_count >= 80:
            wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳商品房信息 - 副本.xlsx')
            run_count = 0
            time.sleep(30)
            #quit()
    wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳商品房信息 - 副本.xlsx')



if __name__ == '__main__':
    writeINFO()