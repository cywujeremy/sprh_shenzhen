import requests
import json
import openpyxl as op
import math
import random

x_pi = 3.14159265358979324 * 3000.0 / 180.0
pi = 3.1415926535897932384626  # π
a = 6378245.0  # 长半轴
ee = 0.00669342162296594323  # 扁率


def getGeoCode(searchLocation):
    api_addr = "https://restapi.amap.com/v3/geocode/geo?address=" + searchLocation + "&city=深圳市&output=json&key=75450f9014e6e0800f5c479088bc0b76"
    req = requests.get(api_addr)
    content = req.content
    sjson = json.loads(content)
    if sjson['status'] == '1':
        try:
            return [sjson['geocodes'][0]['formatted_address'], sjson['geocodes'][0]['location']]
        except IndexError and TypeError:
            print(searchLocation + '-查询不到坐标')
            return ['error', 'error']
    else:
        return ['error', 'error']


def writeGeoCode():
    wb = op.load_workbook('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳市医院信息.xlsx')
    ws = wb['INFO']
    i = 2
    k = ws.max_row
    while i <= ws.max_row:
        try:
            poiInfo = getGeoCode(ws['A' + str(i)].value)
            ws['H' + str(i)].value = poiInfo[0]
            ws['I' + str(i)].value = poiInfo[1]
            print('查询成功！' + ws['A' + str(i)].value)
        except Exception:
            print('查询失败！' + ws['A' + str(i)].value)
            ws['H' + str(i)].value = 'Error'
            ws['I' + str(i)].value = 'Error'
        i += 1
    wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳市医院信息.xlsx')


def gcj02towgs84(lng, lat):
    """
    GCJ02(火星坐标系)转GPS84
    :param lng:火星坐标系的经度
    :param lat:火星坐标系纬度
    :return:
    """

    dlat = transformlat(lng - 105.0, lat - 35.0)
    dlng = transformlng(lng - 105.0, lat - 35.0)
    radlat = lat / 180.0 * pi
    magic = math.sin(radlat)
    magic = 1 - ee * magic * magic
    sqrtmagic = math.sqrt(magic)
    dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrtmagic) * pi)
    dlng = (dlng * 180.0) / (a / sqrtmagic * math.cos(radlat) * pi)
    mglat = lat + dlat
    mglng = lng + dlng
    return [lng * 2 - mglng, lat * 2 - mglat]


def transformlat(lng, lat):
    ret = -100.0 + 2.0 * lng + 3.0 * lat + 0.2 * lat * lat + \
          0.1 * lng * lat + 0.2 * math.sqrt(math.fabs(lng))
    ret += (20.0 * math.sin(6.0 * lng * pi) + 20.0 *
            math.sin(2.0 * lng * pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lat * pi) + 40.0 *
            math.sin(lat / 3.0 * pi)) * 2.0 / 3.0
    ret += (160.0 * math.sin(lat / 12.0 * pi) + 320 *
            math.sin(lat * pi / 30.0)) * 2.0 / 3.0
    return ret


def transformlng(lng, lat):
    ret = 300.0 + lng + 2.0 * lat + 0.1 * lng * lng + \
          0.1 * lng * lat + 0.1 * math.sqrt(math.fabs(lng))
    ret += (20.0 * math.sin(6.0 * lng * pi) + 20.0 *
            math.sin(2.0 * lng * pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lng * pi) + 40.0 *
            math.sin(lng / 3.0 * pi)) * 2.0 / 3.0
    ret += (150.0 * math.sin(lng / 12.0 * pi) + 300.0 *
            math.sin(lng / 30.0 * pi)) * 2.0 / 3.0
    return ret


def coordTransform():
    wb = op.load_workbook('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2020.02深圳/深圳市地铁站信息.xlsx')
    ws = wb['Sheet1']
    i = 2
    while i <= ws.max_row:
        wgscoord = gcj02towgs84(ws['J' + str(i)].value, ws['K' + str(i)].value)
        ws['L'+str(i)] = wgscoord[0]
        ws['M'+str(i)] = wgscoord[1]
        i += 1
    wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2020.02深圳/深圳市地铁站信息.xlsx')



if __name__ == '__main__':
    coordTransform()
