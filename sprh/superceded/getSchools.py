import requests
from bs4 import BeautifulSoup
import openpyxl as op
import re


def htmlIntoSoup(url):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36'}
    response = requests.get(url=url, headers=headers)
    response.encoding = 'GB2312'
    processedtexts = response.text.replace('\n', '')
    soup = BeautifulSoup(processedtexts, 'lxml')
    return soup


def getUrl(soup):
    distlist = soup.find(class_='zx_ml_list').ul.contents
    schools = []
    for item in distlist:
        if item != ' ':
            schools.append(item.find(text='机构概况').parent['href'][2:])
    return schools


def getInfo(soup):
    schName = soup.find(text='学校类别').parent.parent.previous_sibling.previous_sibling.text
    schType = soup.find(text='学校类别').parent.next_sibling.next_sibling.text
    schSuper = soup.find(text='上级主管部门').parent.next_sibling.next_sibling.text
    schAttr = soup.find(text='学校性质').parent.next_sibling.next_sibling.text
    schDist = soup.find(text='所在区').parent.next_sibling.next_sibling.text
    schAdd = soup.find(text='详细地址 ').parent.next_sibling.next_sibling.text
    return [schName, schType, schSuper, schAttr, schDist, schAdd]


def writeDistUrl():
    wb = op.load_workbook('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳市学校信息.xlsx')
    ws1 = wb['District_URLs']
    ws2 = wb['School_URLs']
    i = 2
    while i <= ws1.max_row:
        schurl = getUrl(htmlIntoSoup(ws1['A' + str(i)].value))
        for item in schurl:
            ws2.append([ws1['A' + str(i)].value + item])
        i += 1
    wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳市学校信息.xlsx')


def writeSchInfo():
    wb = op.load_workbook('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳市学校信息.xlsx')
    ws1 = wb['School_URLs']
    ws2 = wb['INFO']
    i = 2
    while i <= ws1.max_row:
        schoolinfo = getInfo(htmlIntoSoup(ws1['A' + str(i)].value))
        ws2.append(schoolinfo)
        i += 1
    wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳市学校信息.xlsx')


if __name__ == '__main__':
    writeSchInfo()
