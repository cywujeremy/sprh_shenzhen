import requests
import json
import openpyxl as op
import math
import random

x_pi = 3.14159265358979324 * 3000.0 / 180.0
pi = 3.1415926535897932384626  # π
a = 6378245.0  # 长半轴
ee = 0.00669342162296594323  # 扁率


def gcj02towgs84(lng, lat):
    """
    GCJ02(火星坐标系)转GPS84
    :param lng:火星坐标系的经度
    :param lat:火星坐标系纬度
    :return:
    """

    dlat = transformlat(lng - 105.0, lat - 35.0)
    dlng = transformlng(lng - 105.0, lat - 35.0)
    radlat = lat / 180.0 * pi
    magic = math.sin(radlat)
    magic = 1 - ee * magic * magic
    sqrtmagic = math.sqrt(magic)
    dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrtmagic) * pi)
    dlng = (dlng * 180.0) / (a / sqrtmagic * math.cos(radlat) * pi)
    mglat = lat + dlat
    mglng = lng + dlng
    return [lng * 2 - mglng, lat * 2 - mglat]


def transformlat(lng, lat):
    ret = -100.0 + 2.0 * lng + 3.0 * lat + 0.2 * lat * lat + \
          0.1 * lng * lat + 0.2 * math.sqrt(math.fabs(lng))
    ret += (20.0 * math.sin(6.0 * lng * pi) + 20.0 *
            math.sin(2.0 * lng * pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lat * pi) + 40.0 *
            math.sin(lat / 3.0 * pi)) * 2.0 / 3.0
    ret += (160.0 * math.sin(lat / 12.0 * pi) + 320 *
            math.sin(lat * pi / 30.0)) * 2.0 / 3.0
    return ret


def transformlng(lng, lat):
    ret = 300.0 + lng + 2.0 * lat + 0.1 * lng * lng + \
          0.1 * lng * lat + 0.1 * math.sqrt(math.fabs(lng))
    ret += (20.0 * math.sin(6.0 * lng * pi) + 20.0 *
            math.sin(2.0 * lng * pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lng * pi) + 40.0 *
            math.sin(lng / 3.0 * pi)) * 2.0 / 3.0
    ret += (150.0 * math.sin(lng / 12.0 * pi) + 300.0 *
            math.sin(lng / 30.0 * pi)) * 2.0 / 3.0
    return ret


def genEstCoord(oriCoord, distance):
    oriLng = float(oriCoord[0:oriCoord.find(',')])
    oriLat = float(oriCoord[oriCoord.find(',') + 1:len(oriCoord)])
    randomArc = random.uniform(0, math.pi * 2)
    lngModify = distance * math.sin(randomArc) * 0.00001141
    latModify = distance * math.cos(randomArc) * 0.00000899
    estLng = oriLng + lngModify
    estLat = oriLat + latModify
    estCoord = str(format(estLng, '.6f')) + ',' + str(format(estLat, '.6f'))
    return estCoord


def locate_current(worksheet1):
    for i in range(1, worksheet1.max_row):
        t = worksheet1['R'][i].value
        if worksheet1['R'][i - 1].value != '' and worksheet1['R'][i].value is None and worksheet1['R'][i].value == \
                worksheet1['R'][i + 1].value and worksheet1['R'][i + 1].value == worksheet1['R'][i + 2].value:
            return i


def locate_current2(worksheet1):
    for i in range(1, worksheet1.max_row):
        t = worksheet1['W'][i].value
        if worksheet1['W'][i - 1].value != '' and worksheet1['W'][i].value is None and worksheet1['W'][i].value == \
                worksheet1['W'][i + 1].value and worksheet1['W'][i + 1].value == worksheet1['W'][i + 2].value:
            return i


def getGeoCode(searchLocation):
    api_addr = "https://restapi.amap.com/v3/geocode/geo?address=" + searchLocation + "&city=深圳市&output=json&key=75450f9014e6e0800f5c479088bc0b76"
    req = requests.get(api_addr)
    content = req.content
    reply = json.loads(content)
    if reply['status'] == '1':
        try:
            return [reply['geocodes'][0]['formatted_address'], reply['geocodes'][0]['location']]
        except IndexError and TypeError:
            print(searchLocation + '-查询不到坐标')
            return ['error', 'error']
    else:
        return ['error', 'error']


def convertCoord():
    wb = op.load_workbook('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳市公园信息.xlsx')
    ws = wb['Sheet1']
    i = 1
    while i <= ws.max_row - 1:
        convertion = gcj02towgs84(ws['J'][i].value, ws['K'][i].value)
        ws['L'][i].value = convertion[0]
        ws['M'][i].value = convertion[1]
        print('转换成功！' + ws['B'][i].value)
        i += 1
    wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳市公园信息.xlsx')


def writeXlsx():
    wb = op.load_workbook('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳小产权房_2.xlsx')
    ws = wb['新房']
    i = locate_current(ws)
    run_control = 0
    while i <= ws.max_row - 1:
        geocode = getGeoCode(ws['O'][i].value)
        geocode[1] = genEstCoord(geocode[1], ws['Q'][i].value)
        ws['R'][i].value = geocode[1]
        ws['S'][i].value = geocode[0]
        print('查询成功！' + ws['E'][i].value)
        run_control += 1
        if run_control >= 600:
            wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳小产权房_2.xlsx')
            quit()
        i += 1
        if i == ws.max_row:
            wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳小产权房_2.xlsx')


if __name__ == '__main__':
    convertCoord()
