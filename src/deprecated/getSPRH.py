from utils.scraper import htmlIntoSoup
import openpyxl as op

def getUrl(soup):
    urlsoup = soup.find(class_='xf_sh_list_b').ul.contents
    sprh = []
    for item in urlsoup:
        if item != '\n':
            try:
                sprh.append(item.find(text='查看地图').parent.parent.previous_sibling.previous_sibling.contents[0]['href'])
            except TypeError:
                sprh.append('error')
    return sprh


def getInfo(soup):
    maintag = soup.find(class_='detail_fl')
    fullname = maintag.find(text='区域/商圈：').parent.parent.parent.previous_sibling.previous_sibling.text
    projname = fullname[fullname.find('【') + 1:fullname.find('】')]
    city = maintag.find(text='区域/商圈：').parent.parent.parent.previous_sibling.previous_sibling.text[:2] + '市'
    subcity = maintag.find(text='区域/商圈：').parent.parent.parent.previous_sibling.previous_sibling.text[2:4] + '区'
    address = maintag.find(text='售楼地址：').parent.next_sibling
    dist = maintag.find(text='区域/商圈：').parent.next_sibling
    subdist = maintag.find(text='环线/片区：').parent.next_sibling
    bldgtype = maintag.find(text='建筑类型：').parent.next_sibling
    sprhtype = maintag.find(text='建筑特色：').parent.next_sibling
    expireterm = maintag.find(text='产权年限：').parent.next_sibling
    projtype = maintag.find(text='项目类型：').parent.next_sibling
    decoration = maintag.find(text='装修情况：').parent.next_sibling
    unitprice = maintag.find(text='参考价格：').parent.next_sibling.next_sibling.text
    fee = maintag.find(text='物 业 费：').parent.next_sibling
    return [projname, city, subcity, address, dist, subdist, bldgtype, sprhtype, expireterm, projtype, decoration,
            unitprice, fee]


def writeDistUrl():
    wb = op.load_workbook('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳市小产权房.xlsx')
    ws1 = wb['URLs']
    ws2 = wb['SPRH_URLs']
    i = 2
    while i <= ws1.max_row:
        sprhurl = getUrl(htmlIntoSoup(ws1['A' + str(i)].value))
        for item in sprhurl:
            ws2.append([item])
        i += 1
    wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳市小产权房.xlsx')


def writeinfo():
    wb = op.load_workbook('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳市小产权房 - 副本.xlsx')
    ws1 = wb['SPRH_URLs']
    ws2 = wb['INFO']
    i = 1
    runcount = 0
    while i <= ws1.max_row:
        sprhinfo = getInfo(htmlIntoSoup(ws1['A' + str(i)].value+'xinxi.html'))
        ws2.append(sprhinfo)
        print(sprhinfo[0]+' 查询成功！')
        runcount += 1
        if runcount >= 50:
            wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳市小产权房 - 副本.xlsx')
            runcount = 0
        i += 1
    wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳市小产权房 - 副本.xlsx')


if __name__ == '__main__':
    writeinfo()
