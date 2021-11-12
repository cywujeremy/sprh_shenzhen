import openpyxl as op

def identify_attr():
    wb = op.load_workbook("D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2020.02深圳/深圳市小产权房.xlsx")
    ws = wb['INFO']
    i = 2
    # print(ws['F2'].value.find('军产房') != -1)
    while i <= ws.max_row:
        if ws['F'+str(i)].value.find('军产房') != -1:
            ws['K'+str(i)] = 3
        elif (ws['F'+str(i)].value.find('村委房') != -1) or (ws['F'+str(i)].value.find('统建房') != -1):
            ws['K' + str(i)] = 2
        else:
            ws['K' + str(i)] = 1
        i += 1
    wb.save("D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2020.02深圳/深圳市小产权房_房屋性质.xlsx")

if __name__ == '__main__':
    identify_attr()