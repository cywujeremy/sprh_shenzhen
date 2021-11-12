from utils.scraper import htmlIntoSoup
import openpyxl as op
import time
import logging

logging.basicConfig(level=logging.INFO)


def getUrl(soup):
    urlsoup = soup.find(class_='li-info').h3.contents
    try:
        sprh = urlsoup[3]['href']
    except TypeError:
        sprh = 'error'
    return sprh


def getINFO(soup):
    
    clip1 = soup.find(class_='comm-title')
    name = clip1.find(class_='sub-hd').previous_sibling
    clip2 = soup.find(class_='basic-parms-mod')
    fee = clip2.find(text='物业费：').parent.next_sibling.next_sibling.text
    
    return name, fee


def writeDistUrl(output_path):
    
    wb = op.load_workbook(output_path)
    ws1 = wb['NAME']
    ws2 = wb['URLs']
    i = ws2.max_row
    if i != 1:
        i = i + 1
    print(i)
    run_count = 0
    err_count = 0
    while i <= ws1.max_row:
        #name = ws1['A' + str(i)].value
        try:
            commurl = getUrl(htmlIntoSoup('https://shenzhen.anjuke.com/community/?kw='+ws1['A' + str(i)].value))
            ws2.append([commurl])
            logging.info(ws1['A' + str(i)].value + '  URL查询成功！')
        except Exception as e:
            # ws2.append(['error'])
            logging.info(ws1['A' + str(i)].value + '  URL查询错误！')
            logging.debug(e)
            err_count += 1
            if err_count >= 3:
                wb.save(output_path)
                time.sleep(60)
                i = i - 4
                err_count = 0
        i += 1
        run_count += 1
        if run_count >= 80:
            wb.save(output_path)
            run_count = 0
            time.sleep(30)
    wb.save(output_path)


def writeINFO(output_path):
    wb = op.load_workbook(output_path)
    ws1 = wb['URLs']
    ws2 = wb['INFO']
    i = ws2.max_row
    if i != 1:
        i = i + 1
    print(i)
    run_count = 0
    err_count = 0
    while i <= ws1.max_row:
        #name = ws1['A' + str(i)].value
        try:
            commurl = getINFO(htmlIntoSoup(ws1['A' + str(i)].value))
            ws2.append(commurl)
            logging.info(commurl[0] + '  信息查询成功！')
        except Exception as e:
            # ws2.append(['error'])
            logging.info(commurl[0] + '  信息查询错误！')
            logging.debug(e)
            err_count += 1
            if err_count >= 3:
                wb.save(output_path)
                time.sleep(60)
                i = i - 4
                err_count = 0
        i += 1
        run_count += 1
        if run_count >= 80:
            wb.save(output_path)
            run_count = 0
            time.sleep(30)
            #quit()
    wb.save(output_path)



if __name__ == '__main__':
    writeINFO()