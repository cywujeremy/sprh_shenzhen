from utils.geocode import gcj02_to_wgs84, getGeoCode
import openpyxl as op

def writeGeoCode():
    wb = op.load_workbook('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳市医院信息.xlsx')
    ws = wb['INFO']
    i = 2
    k = ws.max_row
    while i <= ws.max_row:
        try:
            poiInfo = getGeoCode(ws['A' + str(i)].value)
            ws['H' + str(i)].value = poiInfo[0]
            ws['I' + str(i)].value = poiInfo[1]
            print('查询成功！' + ws['A' + str(i)].value)
        except Exception:
            print('查询失败！' + ws['A' + str(i)].value)
            ws['H' + str(i)].value = 'Error'
            ws['I' + str(i)].value = 'Error'
        i += 1
    wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2019.10 深圳/深圳市医院信息.xlsx')

def coordTransform():
    wb = op.load_workbook('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2020.02深圳/深圳市地铁站信息.xlsx')
    ws = wb['Sheet1']
    i = 2
    while i <= ws.max_row:
        wgscoord = gcj02_to_wgs84(ws['J' + str(i)].value, ws['K' + str(i)].value)
        ws['L'+str(i)] = wgscoord[0]
        ws['M'+str(i)] = wgscoord[1]
        i += 1
    wb.save('D:/OneDriveLocal/OneDrive/学习/毕业/毕业论文/2020.02深圳/深圳市地铁站信息.xlsx')

if __name__ == '__main__':
    coordTransform()
